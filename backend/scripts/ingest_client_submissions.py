"""
Ingest a client's internal submission tracker into public.client_submissions.

Today's only supported source: TAS Europrojects' Excel tracker.
    data/training/TAS Project Management tracker new v.xlsx

Usage:
    python3.12 backend/scripts/ingest_client_submissions.py tas-europrojects \
        --xlsx "data/training/TAS Project Management tracker new v.xlsx" \
        --sheet "List of Submissions 2025-2026"

The script is idempotent: it deletes existing rows for the slug (source matching
the file) before re-inserting. Hard-coded mapping of the TAS columns; future
slugs will need their own column map.
"""
from __future__ import annotations
import argparse
import logging
import sys
import warnings
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(_PROJECT_ROOT / ".env")

from sqlalchemy import text  # noqa: E402
from backend.core.database import SessionLocal  # noqa: E402

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)


def _parse_outcome(raw: str) -> str:
    """Map Excel outcome string -> our enum."""
    s = (raw or "").strip().lower()
    if "not awarded" in s or "not submitted" in s:
        return "not_awarded"
    if "awarded" in s:
        return "awarded"
    if "cancel" in s:
        return "cancelled"
    if "withdraw" in s:
        return "withdrawn"
    return "pending"


def _parse_num(raw):
    if raw is None:
        return None
    s = str(raw).replace(",", "").replace("€", "").replace("$", "").strip()
    s = s.replace(" ", "").replace(" ", "")
    if s in ("", "/", "n/a", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(raw):
    if raw is None:
        return None
    import datetime
    if isinstance(raw, (datetime.date, datetime.datetime)):
        return raw if isinstance(raw, datetime.date) and not isinstance(raw, datetime.datetime) else raw.date()
    s = str(raw).strip()
    if not s or s.lower() in ("/", "n/a"):
        return None
    # Many sheet cells render as "2026-04-01 00:00:00" strings after openpyxl
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%b-%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _classify_ca(raw: str) -> tuple[str | None, str | None, str | None]:
    """Returns (contracting_authority, programme, lot)."""
    s = (raw or "").strip()
    if not s:
        return None, None, None
    upper = s.upper()
    lot = None
    if "LOT 1" in upper: lot = "Lot 1"
    elif "LOT 4" in upper: lot = "Lot 4"
    elif "LOT 5" in upper: lot = "Lot 5"
    elif "LOT 7" in upper: lot = "Lot 7"
    elif "LOT 8" in upper: lot = "Lot 8"
    elif "LOT 2" in upper: lot = "Lot 2"

    programme = None
    if "SEA 2023" in upper: programme = "SEA 2023"
    elif "EIBAS 2022" in upper: programme = "EIBAS 2022"
    elif "SIEA 2018" in upper: programme = "SIEA 2018"
    elif "DG MOVE" in upper: programme = "DG MOVE FWC"
    elif "TRANSPORT COMMUNITY" in upper: programme = "Transport Community"

    if "EIBAS" in upper or "EIB" in upper: ca = "EIB"
    elif "EC " in upper or upper.startswith("EC") or "COMMISSION" in upper: ca = "European Commission"
    elif "WORLD BANK" in upper or upper.strip() == "WB": ca = "World Bank"
    elif "ADB" in upper: ca = "Asian Development Bank"
    elif "AIIB" in upper: ca = "AIIB"
    elif "IDB" in upper: ca = "Inter-American Development Bank"
    elif "TRANSPORT COMMUNITY" in upper: ca = "Transport Community"
    else: ca = s if s else None
    return ca, programme, lot


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("slug")
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default="List of Submissions 2025-2026")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        log.error("Sheet %r not in workbook. Available: %s", args.sheet, wb.sheetnames)
        sys.exit(2)
    ws = wb[args.sheet]

    HEADERS = ["#","project","ca_raw","country","submission_date","status_text",
               "leading_partner","outcome_text","comments","budget","client_margin"]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row[:len(HEADERS)])
        if not vals[1] or not str(vals[1]).strip():
            continue
        r = {HEADERS[i]: (str(vals[i]).strip() if vals[i] is not None else "") for i in range(len(HEADERS))}
        ca, programme, lot = _classify_ca(r["ca_raw"])
        rec = {
            "slug": args.slug,
            "external_ref": r["#"] or None,
            "project_title": r["project"][:1000],
            "contracting_authority": ca,
            "country": r["country"][:200] if r["country"] else None,
            "region": None,
            "programme": programme,
            "lot": lot,
            "sector": None,
            "submission_date": _parse_date(r["submission_date"]),
            "submission_status": r["status_text"][:200] if r["status_text"] else None,
            "leading_partner": (r["leading_partner"].split("\n")[0].split("(")[0].strip()[:200]) if r["leading_partner"] else None,
            "is_lead": "tas" in (r["leading_partner"] or "").lower().split("\n")[0].split("(")[0],
            "outcome": _parse_outcome(r["outcome_text"]),
            "comments": (r["comments"] or "")[:4000] or None,
            "budget_eur": _parse_num(r["budget"]),
            "client_margin_eur": _parse_num(r["client_margin"]),
            "source": "excel",
            "source_file": args.xlsx.rsplit("/", 1)[-1],
            "raw_payload": None,
            "is_test": False,
        }
        rows.append(rec)

    log.info("Parsed %d submissions for slug=%s", len(rows), args.slug)

    if args.dry_run:
        from collections import Counter
        ct = Counter(r["outcome"] for r in rows)
        log.info("Outcome distribution: %s", dict(ct))
        for r in rows[:3]:
            log.info("  e.g. %s", {k: v for k, v in r.items() if k != "raw_payload"})
        return

    db = SessionLocal()
    try:
        db.execute(text("DELETE FROM client_submissions WHERE slug=:slug AND source='excel' AND source_file=:src"),
                   {"slug": args.slug, "src": args.xlsx.rsplit("/", 1)[-1]})
        for r in rows:
            db.execute(text("""
              INSERT INTO client_submissions
                (slug, external_ref, project_title, contracting_authority, country, region,
                 programme, lot, sector, submission_date, submission_status,
                 leading_partner, is_lead, outcome, comments, budget_eur,
                 client_margin_eur, source, source_file, is_test)
              VALUES
                (:slug, :external_ref, :project_title, :contracting_authority, :country, :region,
                 :programme, :lot, :sector, :submission_date, :submission_status,
                 :leading_partner, :is_lead, :outcome, :comments, :budget_eur,
                 :client_margin_eur, :source, :source_file, :is_test);
            """), r)
        db.commit()
        from sqlalchemy import text as _t
        ct = db.execute(_t("SELECT outcome, COUNT(*) FROM client_submissions WHERE slug=:slug GROUP BY outcome"),
                        {"slug": args.slug}).fetchall()
        for o, n in ct:
            log.info("  %s: %d", o, n)
    finally:
        db.close()
    log.info("Done.")


if __name__ == "__main__":
    main()
