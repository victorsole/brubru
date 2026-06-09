"""
EMA — European Medicines Agency (api_health.md). /api/v2/ema.

EMA runs the EU OpenEuropa / BCL Drupal distribution (ema.europa.eu). Source map
(verified):
  - news   : /en/news — bcl cards; each item is a /en/news/<slug> link inside a
             .card-body whose text carries the date ("3 June 2026"). ?page=N paginated.
  - events : /en/events/upcoming-events — cards with a <time datetime> and an
             /en/events/<slug> link (excluding the upcoming/past nav links).

EMA's /news-and-events/publications page renders its list client-side (EAMS) with
no server-side items, so only news + events are surfaced. No LLM is used.
"""
from __future__ import annotations

import re
import time
from datetime import datetime, timezone

from bs4 import BeautifulSoup

from services.scrapers.economy_common import (
    Item, clean, norm_url, http_get, fetch_detail, parse_listing_date, _iso_dt,
)

_BASE = "https://www.ema.europa.eu"
_NEWS_SLUG = re.compile(r"^/en/news/[a-z0-9]")
_EVENT_NAV = {"upcoming-events", "past-events", "events"}


def _parse_news(html: str):
    main = BeautifulSoup(html, "html.parser").select_one("main") or BeautifulSoup(html, "html.parser")
    out = []
    seen = set()
    for a in main.find_all("a", href=True):
        href = a["href"].split("?")[0]
        if not _NEWS_SLUG.match(href) or not a.get_text(strip=True):
            continue
        url = norm_url(_BASE + href)
        if url in seen:
            continue
        title = clean(a.get_text(" ", strip=True))
        if not title or len(title) < 12:
            continue
        seen.add(url)
        node = a
        doc_dt = None
        for _ in range(5):
            node = node.parent
            if node is None:
                break
            doc_dt = parse_listing_date(node.get_text(" ", strip=True))
            if doc_dt:
                break
        out.append((url, title, doc_dt))
    return out


def _parse_events(html: str):
    main = BeautifulSoup(html, "html.parser").select_one("main") or BeautifulSoup(html, "html.parser")
    out = []
    seen = set()
    for t in main.select("time[datetime]"):
        node, link = t, None
        for _ in range(5):
            node = node.parent
            if node is None:
                break
            for a in node.find_all("a", href=True):
                href = a["href"].split("?")[0]
                slug = href.rstrip("/").split("/")[-1]
                if "/en/events/" in href and slug not in _EVENT_NAV and a.get_text(strip=True):
                    link = a
                    break
            if link:
                break
        if not link:
            continue
        url = norm_url(_BASE + link["href"].split("?")[0])
        if url in seen:
            continue
        seen.add(url)
        title = clean(link.get_text(" ", strip=True))
        if not title:
            continue
        out.append((url, title, _iso_dt(t.get("datetime")) or parse_listing_date(t.get_text(" ", strip=True))))
    return out


def _scrape(item_type, listing, parser, *, fetch_bodies, max_pages):
    items, seen = [], set()
    now = datetime.now(timezone.utc)
    for page in range(max_pages):
        sep = "&" if "?" in listing else "?"
        url = listing if page == 0 else f"{listing}{sep}page={page}"
        r = http_get(url)
        if r is None:
            break
        rows = parser(r.text)
        new = 0
        for u, title, doc_dt in rows:
            if u in seen:
                continue
            seen.add(u)
            new += 1
            items.append(Item(body_code="ema", item_type=item_type, title=title,
                              public_url=u, document_date=doc_dt, creation_date=now,
                              source_kind="html", guid=u))
        if new == 0:
            break
    if fetch_bodies:
        # EMA returns HTTP 429 on bursts; space the detail fetches and retry the
        # transient throttle/timeout failures so bodies extract cleanly.
        for it in items:
            body_txt = body_html = None
            kind = "unreachable"
            for attempt in range(4):
                body_txt, body_html, kind = fetch_detail(it.public_url)
                if kind != "unreachable":
                    break
                time.sleep(3.0 * (attempt + 1))
            it.body_txt, it.body_html = body_txt, body_html
            if kind == "pdf":
                it.source_kind = "pdf"
            time.sleep(0.6)
    return items


def ingest_ema_news(*, fetch_bodies: bool = True, max_pages: int = 6) -> list[Item]:
    return _scrape("news", f"{_BASE}/en/news", _parse_news, fetch_bodies=fetch_bodies, max_pages=max_pages)


# --- medicines / EPARs (structured dataset) --------------------------------
# EMA's published assessment reports (EPARs) are not exposed as a clean listing
# on the website, but the full register is downloadable as an .xlsx. Each row is
# one medicine (name, EMA product number, status, INN, therapeutic area, key
# dates, EPAR URL) — the row IS the content, so no per-medicine HTTP fetch.
_MEDICINES_XLSX = f"{_BASE}/en/documents/report/medicines-output-medicines-report_en.xlsx"
_MED_FIELDS = [  # (column header substring, label) — rendered into the body
    ("Category", "Category"), ("EMA product number", "EMA product number"),
    ("Medicine status", "Status"), ("International non-proprietary name", "INN"),
    ("Active substance", "Active substance"), ("Therapeutic area", "Therapeutic area"),
    ("Therapeutic indication", "Therapeutic indication"),
    ("Marketing authorisation developer", "Marketing-authorisation holder"),
    ("ATC code", "ATC code"), ("Orphan medicine", "Orphan"),
    ("European Commission decision date", "EC decision date"),
    ("Marketing authorisation date", "Marketing-authorisation date"),
    ("First published date", "First published"), ("Last updated date", "Last updated"),
]


def _ema_date(val) -> datetime | None:
    from datetime import date as _date
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc)
    if isinstance(val, _date):
        return datetime(val.year, val.month, val.day, tzinfo=timezone.utc)
    return parse_listing_date(str(val)) if val else None


def ingest_ema_medicines(*, fetch_bodies: bool = True, **_) -> list[Item]:
    import io
    from openpyxl import load_workbook
    r = http_get(_MEDICINES_XLSX)
    if r is None:
        return []
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # locate the header row (the one containing "Name of medicine")
    hdr_idx = next((i for i, row in enumerate(rows)
                    if any(isinstance(c, str) and "Name of medicine" in c for c in row)), None)
    if hdr_idx is None:
        return []
    header = [str(c).replace("\n", " ").strip() if c is not None else "" for c in rows[hdr_idx]]

    def col(substr):
        return next((i for i, h in enumerate(header) if substr.lower() in h.lower()), None)

    i_name, i_url = col("Name of medicine"), col("Medicine URL")
    i_pub, i_upd = col("First published date"), col("Last updated date")
    field_cols = [(label, col(sub)) for sub, label in _MED_FIELDS]
    now = datetime.now(timezone.utc)
    items: list[Item] = []
    seen: set[str] = set()
    for row in rows[hdr_idx + 1:]:
        if i_name is None or i_name >= len(row) or not row[i_name]:
            continue
        name = clean(str(row[i_name]).strip())
        url = (str(row[i_url]).strip() if i_url is not None and row[i_url] else None)
        guid = url or name
        if not name or guid in seen:
            continue
        seen.add(guid)
        doc_dt = _ema_date(row[i_pub]) if (i_pub is not None and i_pub < len(row)) else None
        if doc_dt is None and i_upd is not None and i_upd < len(row):
            doc_dt = _ema_date(row[i_upd])
        lines = []
        for label, ci in field_cols:
            if ci is not None and ci < len(row) and row[ci] not in (None, ""):
                lines.append(f"{label}: {str(row[ci]).strip()}")
        body_txt = clean("\n".join(lines)) or None
        body_html = clean("<ul>" + "".join(f"<li>{l}</li>" for l in lines) + "</ul>") if lines else None
        summary = clean(" | ".join(lines[:4])) if lines else None
        items.append(Item(
            body_code="ema", item_type="medicine", title=name,
            public_url=norm_url(url) if url else f"{_BASE}/en/medicines",
            summary=summary, body_txt=body_txt, body_html=body_html,
            document_date=doc_dt, creation_date=now, source_kind="dataset", guid=guid,
        ))
    return items


def ingest_ema_events(*, fetch_bodies: bool = True, max_pages: int = 4) -> list[Item]:
    return _scrape("event", f"{_BASE}/en/events/upcoming-events", _parse_events,
                   fetch_bodies=fetch_bodies, max_pages=max_pages)
