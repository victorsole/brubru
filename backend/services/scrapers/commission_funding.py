"""
EU Financial Transparency System (FTS) — the database behind
/api/v2/commission/eu-funding-recipients.

Source: the FTS annual datasets on the EU Open Data Portal, published as one
Excel file per year at
  https://ec.europa.eu/budget/financial-transparency-system/download/<year>_FTS_dataset_en.xlsx
Each row is one beneficiary's share of a legal commitment (grant, procurement
contract, prize, etc.) under EU direct management. We ingest the most recent
complete year (auto-discovered). The row IS the content — no per-row fetch, no LLM.

The legal-commitment reference repeats across rows (one commitment can have many
beneficiaries), so the unique key is a hash of the salient fields.
"""
from __future__ import annotations

import hashlib
import io
from datetime import datetime, timezone

import requests
# openpyxl is imported lazily inside the one function that needs it (below), so a
# missing optional Excel dependency can never break importing this module -- which
# sync_economy.py imports at top level for EVERY body. A hard top-level import here
# froze ALL economy ingestion on prod ~25 Jun to 10 Jul 2026 (openpyxl not in
# requirements.txt at the time). Fixed 10 Jul: added to requirements + made lazy.

from services.scrapers.economy_common import Item, clean, _iso_dt

_BASE = "https://ec.europa.eu/budget/financial-transparency-system/download"
_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36")


def _fetch_latest() -> tuple[int, bytes] | None:
    """The most recent year whose Excel file actually exists. The server returns
    a 200 HTML page for not-yet-published years, so we validate the XLSX (zip)
    magic bytes rather than trusting the status code."""
    now = datetime.now(timezone.utc).year
    for y in range(now, 2006, -1):
        try:
            r = requests.get(f"{_BASE}/{y}_FTS_dataset_en.xlsx",
                             headers={"User-Agent": _UA}, timeout=240)
            if r.status_code == 200 and r.content[:2] == b"PK":
                return y, r.content
        except requests.RequestException:
            continue
    return None


def _col(idx: dict, *subs: str):
    for sub in subs:
        for h, i in idx.items():
            if h and sub.lower() in h.lower():
                return i
    return None


def ingest_eu_funding_recipients(*, fetch_bodies: bool = True, year: int | None = None,
                                 **_) -> list[Item]:
    if year:
        try:
            raw = requests.get(f"{_BASE}/{year}_FTS_dataset_en.xlsx",
                               headers={"User-Agent": _UA}, timeout=240).content
        except requests.RequestException:
            return []
    else:
        found = _fetch_latest()
        if not found:
            return []
        year, raw = found
    try:
        from openpyxl import load_workbook  # lazy import (see module-top note)
        ws = load_workbook(io.BytesIO(raw), read_only=True).active
    except (ValueError, OSError, ImportError):
        return []
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {h: i for i, h in enumerate(header)}
    c_name = _col(idx, "Name of benefic")
    c_lc = _col(idx, "Reference of the Leg")
    c_bref = _col(idx, "Reference (Budget)")
    c_amt = _col(idx, "Commitment  total", "Commitment total")
    c_prog = _col(idx, "Programme name")
    c_subj = _col(idx, "Subject of grant")
    c_cty = _col(idx, "Beneficiary country")
    c_fund = _col(idx, "Funding type")
    c_mgmt = _col(idx, "Management type")
    c_dept = _col(idx, "Responsible depart")
    c_bln = _col(idx, "Budget line name")
    c_start = _col(idx, "Project start")
    if c_name is None:
        return []

    def g(row, i):
        if i is None or i >= len(row):
            return None
        v = row[i]
        return v if v not in (None, "", "-", "*****") else None

    now = datetime.now(timezone.utc)
    items: list[Item] = []
    seen: set[str] = set()
    for row in rows:
        name = clean(g(row, c_name))
        if not name:
            continue
        lc = g(row, c_lc) or ""
        bref = g(row, c_bref) or ""
        amount = g(row, c_amt)
        key = hashlib.md5(f"{lc}|{bref}|{name}|{amount}".encode("utf-8")).hexdigest()[:12]
        if key in seen:
            continue
        seen.add(key)
        url_row = f"https://ec.europa.eu/budget/financial-transparency-system/index.html#{year}-{key}"
        prog = clean(g(row, c_prog))
        country = clean(g(row, c_cty))
        fund = clean(g(row, c_fund))
        subj = clean(g(row, c_subj))
        start = g(row, c_start)
        doc_dt = _iso_dt(str(start)[:10]) if start else _iso_dt(f"{year}-01-01")
        amt_str = None
        if amount is not None:
            try:
                amt_str = f"EUR {float(amount):,.0f}"
            except (TypeError, ValueError):
                amt_str = str(amount)
        lines = [
            f"Beneficiary: {name}",
            f"EU commitment (total): {amt_str}" if amt_str else "",
            f"Funding type: {fund}" if fund else "",
            f"Programme: {prog}" if prog else "",
            f"Subject: {subj}" if subj else "",
            f"Responsible department: {clean(g(row, c_dept))}" if g(row, c_dept) else "",
            f"Budget line: {clean(g(row, c_bln))}" if g(row, c_bln) else "",
            f"Management type: {clean(g(row, c_mgmt))}" if g(row, c_mgmt) else "",
            f"Beneficiary country: {country}" if country else "",
            f"Legal commitment ref: {lc}" if lc else "",
            f"Year: {year}",
        ]
        lines = [l for l in lines if l]
        items.append(Item(
            body_code="commission", item_type="funding_recipient", title=name,
            public_url=url_row,
            summary=clean(" | ".join(x for x in [name, amt_str or "", prog or "", country or ""] if x)),
            body_txt=clean("\n".join(lines)),
            body_html=clean("<ul>" + "".join(f"<li>{l}</li>" for l in lines) + "</ul>"),
            document_date=doc_dt, creation_date=now, source_kind="fts", guid=key))
    return items
