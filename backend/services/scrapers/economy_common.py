"""
Shared primitives for the economy & finance scrapers (ECB, the EU financial
institutions, ESM). Each body's module (economy_ecb.py, economy_eba.py, ...)
imports these and adds its own source map + extraction strategy.

No LLM is used anywhere in economy ingestion.
"""
from __future__ import annotations

import io
import re
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone

import requests
from bs4 import BeautifulSoup

_UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 BrubruBot/1.0"
_TIMEOUT = 30
_BODY_CAP = 800_000  # chars; working-paper PDFs can be large


@dataclass
class Item:
    body_code: str
    item_type: str            # news | publication | event | legal
    title: str
    public_url: str
    summary: str | None = None
    body_txt: str | None = None
    body_html: str | None = None
    document_date: datetime | None = None
    creation_date: datetime | None = None
    source_kind: str | None = None
    guid: str | None = None
    extras: dict = field(default_factory=dict)


_CTRL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")  # C0 controls except \t \n \r


def clean(s: str | None) -> str | None:
    """Strip NUL + other C0 control chars Postgres TEXT rejects (PDF extraction emits them)."""
    if not s:
        return s
    return _CTRL_RE.sub("", s)


def norm_url(url: str) -> str:
    """Collapse doubled slashes after the host (some feeds emit ecb.europa.eu//press/...)."""
    if not url:
        return url
    url = url.strip()
    m = re.match(r"^(https?://[^/]+)(/.*)$", url)
    if not m:
        return url
    return m.group(1) + re.sub(r"/{2,}", "/", m.group(2))


def http_get(url: str) -> requests.Response | None:
    try:
        r = requests.get(url, headers={"User-Agent": _UA}, timeout=_TIMEOUT, allow_redirects=True)
        if r.status_code == 200:
            return r
    except requests.RequestException:
        return None
    return None


def to_dt(struct) -> datetime | None:
    if not struct:
        return None
    try:
        return datetime(*struct[:6], tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return None


def extract_html(html: str) -> tuple[str | None, str | None]:
    """(body_txt, body_html) from a server-rendered detail page."""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "nav", "header", "footer", "aside", "form"]):
        tag.decompose()
    node = soup.find("main") or soup.find("article") or soup.body or soup
    body_html = clean(str(node)[:_BODY_CAP])
    body_txt = clean(node.get_text("\n", strip=True)[:_BODY_CAP])
    return (body_txt or None), (body_html or None)


def extract_pdf(content: bytes) -> str | None:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        parts, total = [], 0
        for page in reader.pages:
            t = page.extract_text() or ""
            parts.append(t)
            total += len(t)
            if total > _BODY_CAP:
                break
        return clean("\n".join(parts).strip()[:_BODY_CAP]) or None
    except Exception:
        return None


def fetch_detail(url: str) -> tuple[str | None, str | None, str]:
    """Return (body_txt, body_html, source_kind) for one item URL.

    HTML -> parsed; PDF -> text-extracted; any other type (Office docs, zip,
    images) -> a link wrapper, never fed to the HTML parser as binary.
    """
    r = http_get(url)
    if r is None:
        return None, None, "unreachable"
    ctype = (r.headers.get("content-type") or "").lower()
    if "application/pdf" in ctype or url.lower().endswith(".pdf"):
        return extract_pdf(r.content), f'<p>PDF document: <a href="{url}">{url}</a></p>', "pdf"
    if "html" not in ctype and "xml" not in ctype:
        # Office docs (.docx/.xlsx), archives, images, etc. — do not parse as HTML.
        kind = (ctype.split(";")[0].split("/")[-1] or "file")[:20] or "file"
        return None, f'<p>Document: <a href="{url}">{url}</a></p>', kind
    body_txt, body_html = extract_html(r.text)
    return body_txt, body_html, "html"


# --- date parsing ----------------------------------------------------------
# Handles: 11/06/2026 ; 29.05.2026 ; "8 June 2026" / "8 JUNE 2026" ; "8 Dec 2026" ;
# ordinals "2nd June 2026" ; ranges "11/06/2026 - 12/06/2026" and "18-19 Nov 2026"
# (-> the start day).
_MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
    "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
_NUM_DATE_RE = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b")
_DOT_DATE_RE = re.compile(r"\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b")
_ORDINAL_RE = re.compile(r"\b(\d{1,2})(?:st|nd|rd|th)\b", re.IGNORECASE)
_NAME_DATE_RE = re.compile(r"\b(\d{1,2})(?:\s*[-–]\s*\d{1,2})?\s+([A-Za-z]{3,9})\.?\s+(\d{4})\b")


def _iso_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def parse_listing_date(text: str) -> datetime | None:
    text = _ORDINAL_RE.sub(r"\1", text or "")    # "2nd June 2026" -> "2 June 2026"
    m = _NUM_DATE_RE.search(text)                # 11/06/2026 (ranges -> first match = start)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(year, month, day, tzinfo=timezone.utc)
        except ValueError:
            return None
    m = _DOT_DATE_RE.search(text)                # 29.05.2026 (day.month.year)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(year, month, day, tzinfo=timezone.utc)
        except ValueError:
            return None
    m = _NAME_DATE_RE.search(text)               # 8 June 2026 / 8 Dec 2026 / 18-19 Nov 2026
    if m:
        month = _MONTHS.get(m.group(2).lower().rstrip("."))
        if not month:
            return None
        try:
            return datetime(int(m.group(3)), month, int(m.group(1)), tzinfo=timezone.utc)
        except ValueError:
            return None
    return None


# --- generic EU ECL (Europa Component Library) listing scraper --------------
# Many europa.eu bodies (EIOPA, AMLA, EPPO, Commission DGs) render listings as
# .ecl-content-item cards with a <time datetime> date, a standalone title link
# and a description. Server-rendered + ?page=N paginated.


def _parse_ecl_cards(html: str, base: str):
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for ci in soup.select(".ecl-content-item"):
        a = ci.select_one(".ecl-content-block__title a[href]") or ci.select_one("a[href]")
        if not a or not a.get("href"):
            continue
        href = a["href"]
        url = norm_url(href if href.startswith("http") else base + href)
        title = clean(a.get_text(" ", strip=True))
        if not title:
            continue
        t = ci.select_one("time[datetime]")
        doc_dt = _iso_dt(t["datetime"]) if (t and t.get("datetime")) else None
        if doc_dt is None:
            # ECL date-block component (events): day / month / year split spans.
            db = ci.select_one(".ecl-date-block")
            if db:
                day = db.select_one(".ecl-date-block__day")
                mon = db.select_one(".ecl-date-block__month")
                yr = db.select_one(".ecl-date-block__year")
                mon_txt = (mon.get("title") or mon.get_text(strip=True)) if mon else ""
                if day and mon_txt and yr:
                    doc_dt = parse_listing_date(
                        f"{day.get_text(strip=True)} {mon_txt} {yr.get_text(strip=True)}")
        if doc_dt is None:
            meta = ci.select_one(".ecl-content-block__primary-meta-item time") or t
            doc_dt = parse_listing_date(meta.get_text(" ", strip=True)) if meta else None
        desc = ci.select_one(".ecl-content-block__description")
        summary = clean(desc.get_text(" ", strip=True)[:1000]) if desc else None
        out.append((url, title, doc_dt, summary))
    return out


def scrape_ecl_listing(body_code: str, item_type: str, listing_urls, base: str,
                       *, fetch_bodies: bool = True, max_pages: int = 6):
    """Page through ECL .ecl-content-item listings and build Items."""
    items = []
    seen: set[str] = set()
    now = datetime.now(timezone.utc)
    for listing in listing_urls:
        for page in range(max_pages):
            sep = "&" if "?" in listing else "?"
            url = listing if page == 0 else f"{listing}{sep}page={page}"
            r = http_get(url)
            if r is None:
                break
            rows = _parse_ecl_cards(r.text, base)
            if not rows:
                break
            new = 0
            for url_i, title, doc_dt, summary in rows:
                if url_i in seen:
                    continue
                seen.add(url_i)
                new += 1
                items.append(Item(body_code=body_code, item_type=item_type, title=title,
                                  public_url=url_i, summary=summary, document_date=doc_dt,
                                  creation_date=now, source_kind="html", guid=url_i))
            if new == 0:
                break
    if fetch_bodies:
        for it in items:
            body_txt, body_html, kind = fetch_detail(it.public_url)
            it.body_txt, it.body_html = body_txt, body_html
            if kind == "pdf":
                it.source_kind = "pdf"
    return items


# --- generic EU ECL file-library scraper -----------------------------------
# ECL document libraries (AMLA, and other europa.eu bodies) render each document
# as an .ecl-file card: a title in .ecl-file__title, type + date in
# .ecl-file__detail-meta-item spans, and a /document/download/<uuid> PDF link.
# Server-rendered + ?page=N paginated.
def _parse_ecl_file_cards(html: str, base: str):
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for f in soup.select(".ecl-file"):
        a = f.select_one("a[href]")
        title_el = f.select_one(".ecl-file__title")
        if not a or not a.get("href") or not title_el:
            continue
        href = a["href"]
        url = norm_url(href if href.startswith("http") else base + href)
        title = clean(title_el.get_text(" ", strip=True))
        if not title:
            continue
        doc_dt = None
        metas = f.select(".ecl-file__detail-meta-item")
        for m in metas:                       # the date meta is the one that parses
            d = parse_listing_date(m.get_text(" ", strip=True))
            if d:
                doc_dt = d
                break
        out.append((url, title, doc_dt))
    return out


def scrape_ecl_file_listing(body_code: str, item_type: str, listing_urls, base: str,
                            *, fetch_bodies: bool = True, max_pages: int = 8):
    """Page through ECL .ecl-file document libraries and build Items (PDFs)."""
    items = []
    seen: set[str] = set()
    now = datetime.now(timezone.utc)
    for listing in listing_urls:
        for page in range(max_pages):
            sep = "&" if "?" in listing else "?"
            url = listing if page == 0 else f"{listing}{sep}page={page}"
            r = http_get(url)
            if r is None:
                break
            rows = _parse_ecl_file_cards(r.text, base)
            if not rows:
                break
            new = 0
            for url_i, title, doc_dt in rows:
                if url_i in seen:
                    continue
                seen.add(url_i)
                new += 1
                items.append(Item(body_code=body_code, item_type=item_type, title=title,
                                  public_url=url_i, document_date=doc_dt, creation_date=now,
                                  source_kind="pdf", guid=url_i))
            if new == 0:
                break
    if fetch_bodies:
        for it in items:
            body_txt, body_html, kind = fetch_detail(it.public_url)
            it.body_txt, it.body_html = body_txt, body_html
            it.source_kind = kind if kind in ("pdf", "html") else it.source_kind
    return items


# --- generic curated-page "topics" snapshotter -----------------------------
# Snapshot a curated list of stable reference / thematic landing pages as topic
# Items: title from the page h1 (or <title>, else the slug) and the body via
# extract_html. Reference pages carry no publication date, so document_date is
# left null. Used across single-body agency folders for their "topics" resource.
def snapshot_topics(body_code: str, base: str, paths, *, fetch_bodies: bool = True,
                    title_max: int = 300, pace: float = 0.3, retries: int = 2,
                    prefer_title: bool = False):
    # prefer_title=True reads the title from <title> before the page h1 — for sites
    # whose h1 is a fixed banner/slogan repeated on every page (e.g. CdT).
    items = []
    now = datetime.now(timezone.utc)
    for path in paths:
        url = base + path if path.startswith("/") else path
        # ec.europa.eu subdomains rate-limit bursts: pace requests and retry the
        # transient blocks (None) a couple of times before giving up on a page.
        r = http_get(url)
        attempt = 0
        while r is None and attempt < retries:
            attempt += 1
            time.sleep(1.5 * attempt)
            r = http_get(url)
        time.sleep(pace)
        if r is None:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        h1 = soup.select_one("main h1, h1")
        title_tag = soup.title.get_text(strip=True) if soup.title else ""
        if prefer_title and title_tag:
            title = clean(title_tag.split(" | ")[0].split(" - ")[0])
        elif h1 and len(h1.get_text(strip=True)) > 2:
            title = clean(h1.get_text(" ", strip=True))
        elif title_tag:
            title = clean(title_tag.split(" | ")[0].split(" - ")[0])
        else:
            title = clean(url.rstrip("/").rsplit("/", 1)[-1].replace("_en", "").replace("-", " ").title())
        body_txt, body_html = (extract_html(r.text) if fetch_bodies else (None, None))
        items.append(Item(body_code=body_code, item_type="topic", title=(title or url)[:title_max],
                          public_url=norm_url(url), creation_date=now, source_kind="html",
                          guid=norm_url(url), body_txt=body_txt, body_html=body_html))
    return items


# --- generic ECB/ESRB date-indexed <dl> listing scraper --------------------
# ECB-CMS bodies (ECB, ESRB) render listings as <dt>(date) + <dd>(title+link).
# Some index pages are server-rendered; others inject the <dl> client-side
# (use_playwright=True). parse_dl_cards handles either rendered HTML.
def parse_dl_cards(html: str, base: str):
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for dd in soup.find_all("dd"):
        a = dd.find("a", href=True)
        if not a:
            continue
        href = a["href"]
        url = norm_url(href if href.startswith("http") else base + href)
        if "/shared/" in url:
            continue
        title = clean(a.get_text(" ", strip=True))
        if not title:
            continue
        dt_node = dd.find_previous_sibling("dt")
        doc_dt = parse_listing_date(dt_node.get_text(" ", strip=True)) if dt_node else None
        out.append((url, title, doc_dt))
    return out


def scrape_dl_listing(body_code: str, item_type: str, pages, base: str, *,
                      use_playwright: bool = False, fetch_bodies: bool = True,
                      settle_ms: int = 6000):
    items = []
    seen: set[str] = set()
    now = datetime.now(timezone.utc)

    def _collect(html: str):
        for url, title, doc_dt in parse_dl_cards(html or "", base):
            if url in seen:
                continue
            seen.add(url)
            items.append(Item(body_code=body_code, item_type=item_type, title=title,
                              public_url=url, document_date=doc_dt, creation_date=now,
                              source_kind="html", guid=url))

    if use_playwright:
        from services.scrapers.waf_browser_fetcher import WafBrowserFetcher
        with WafBrowserFetcher(settle_ms=settle_ms) as f:
            for page in pages:
                res = f.fetch(page, strip_chrome=False)
                _collect(res.html)
    else:
        for page in pages:
            r = http_get(page)
            if r is not None:
                _collect(r.text)

    if fetch_bodies:
        for it in items:
            body_txt, body_html, kind = fetch_detail(it.public_url)
            it.body_txt, it.body_html = body_txt, body_html
            if kind == "pdf":
                it.source_kind = "pdf"
    return items


# --- generic .xlsx dataset ingestor ----------------------------------------
# Many EU agencies publish a register/database as a downloadable .xlsx whose rows
# are entities (medicines, shortages, varieties, ...). Each row IS the content,
# so no per-row HTTP fetch. Header is auto-detected (the row with the most cells).
def _xlsx_date(val):
    from datetime import date as _date
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc)
    if isinstance(val, _date):
        return datetime(val.year, val.month, val.day, tzinfo=timezone.utc)
    return parse_listing_date(str(val)) if val else None


def ingest_xlsx_dataset(xlsx_url, body_code, item_type, *,
                        title_cols, url_cols=(), date_cols=(), max_body_cols=40,
                        sheet=None, record_url=None, timeout=180):
    """Build Items from a downloadable .xlsx register.

    title_cols / url_cols / date_cols are ordered lists of header substrings;
    the first match (that is non-empty for the row, for the title) wins.

    sheet      : worksheet name to read (default: the active sheet).
    record_url : canonical dataset URL; when a row has no own URL, the item's
                 public_url is synthesised as ``{record_url}#row=<n>`` so the
                 (body_code, item_type, public_url) UNIQUE constraint holds.
    """
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        resp = requests.get(xlsx_url, headers={"User-Agent": _UA}, timeout=timeout,
                            allow_redirects=True)
        if resp.status_code != 200:
            return []
    except requests.RequestException:
        return []
    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    scan = min(15, len(rows))
    hdr_idx = max(range(scan), key=lambda i: sum(1 for c in rows[i] if c is not None))
    header = [str(c).replace("\n", " ").strip() if c is not None else "" for c in rows[hdr_idx]]

    def cols_for(cands):
        out = []
        for cand in cands:
            for i, h in enumerate(header):
                if cand.lower() in h.lower() and i not in out:
                    out.append(i)
        return out

    title_idxs = cols_for(title_cols)
    url_idxs = cols_for(url_cols)
    date_idxs = cols_for(list(date_cols) + ["date"])
    now = datetime.now(timezone.utc)
    items, seen = [], set()
    for rownum, row in enumerate(rows[hdr_idx + 1:]):
        title = None
        for ti in title_idxs:
            if ti < len(row) and row[ti] not in (None, ""):
                title = clean(str(row[ti]).strip())
                if title:
                    break
        if not title:
            # fallback: first non-empty text cell whose header is not a url/date
            for i, h in enumerate(header):
                if i < len(row) and isinstance(row[i], str) and row[i].strip() \
                        and "url" not in h.lower() and "date" not in h.lower():
                    title = clean(row[i].strip())
                    if title:
                        break
        if not title:
            continue
        url = None
        for ui in url_idxs:
            if ui < len(row) and row[ui]:
                url = str(row[ui]).strip()
                break
        if not url and record_url:
            # no per-row URL: synthesise a unique one so the UNIQUE constraint holds
            url = f"{record_url}#row={rownum + 1}"
        guid = url or f"{item_type}:{title}:{rownum}"
        if guid in seen:
            continue
        seen.add(guid)
        doc_dt = None
        for di in date_idxs:
            if di < len(row):
                doc_dt = _xlsx_date(row[di])
                if doc_dt:
                    break
        lines = []
        for i, h in enumerate(header[:max_body_cols]):
            if h and i < len(row) and row[i] not in (None, "") and "url" not in h.lower():
                lines.append(f"{h}: {str(row[i]).strip()}")
        body_txt = clean("\n".join(lines)) or None
        body_html = clean("<ul>" + "".join(f"<li>{l}</li>" for l in lines) + "</ul>") if lines else None
        summary = clean(" | ".join(lines[:4])) if lines else None
        items.append(Item(body_code=body_code, item_type=item_type, title=title,
                          public_url=norm_url(url) if url else None,
                          summary=summary, body_txt=body_txt, body_html=body_html,
                          document_date=doc_dt, creation_date=now,
                          source_kind="dataset", guid=guid))
    return items
